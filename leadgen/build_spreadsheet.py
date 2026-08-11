"""Build the NE Wisconsin mortgage-company prospect spreadsheet.

Usage:
    export GOOGLE_MAPS_API_KEY=...
    python3 leadgen/build_spreadsheet.py
"""

import concurrent.futures
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import places
import website_audit

OUT = "NE_Wisconsin_Mortgage_Prospects.xlsx"

# Row colors by opportunity tier.
FILL_HOT = PatternFill("solid", fgColor="FFC7CE")
FILL_WARM = PatternFill("solid", fgColor="FFEB9C")
FILL_COOL = PatternFill("solid", fgColor="D9EAD3")

HEADERS = [
    "Priority", "Business Name", "City", "Website Status", "Website URL",
    "Google Reviews", "Rating", "Phone", "Address", "Why They Need Us",
    "Google Maps Link", "Opportunity Score",
]


def excluded(name):
    low = name.lower()
    return any(pat in low for pat in config.EXCLUDE_NAME_PATTERNS)


def city_of(address):
    """Pull the city out of a formatted US address."""
    m = re.search(r",\s*([^,]+),\s*WI", address or "")
    return m.group(1).strip() if m else ""


def opportunity_score(site_score, reviews):
    """Website condition dominates; review count is a viability multiplier.

    A dead site at a 200-review shop is worth far more than a dead site at a
    16-review shop, so reviews scale the result rather than just adding to it.
    """
    viability = min(reviews / float(config.MAX_REVIEWS), 1.0)  # 0..1
    return round(site_score * (0.70 + 0.30 * viability), 1)


def main():
    # Fail fast on a missing/bad key: collect() tolerates per-query errors, so
    # without this an unset key just yields an empty spreadsheet.
    places.verify_credentials()

    print("1/4  Searching Google Places across the territory...")
    raw = places.collect(config.TARGET_TOWNS, config.SEARCH_TERMS)
    print(f"     {len(raw)} unique businesses found\n")

    print("2/4  Applying filters (review window, brand exclusions)...")
    candidates = []
    for p in raw:
        name = (p.get("displayName") or {}).get("text", "")
        reviews = p.get("userRatingCount") or 0
        if not name or excluded(name):
            continue
        if p.get("businessStatus") not in (None, "OPERATIONAL"):
            continue
        if not (config.MIN_REVIEWS <= reviews <= config.MAX_REVIEWS):
            continue
        candidates.append(p)
    print(f"     {len(candidates)} passed the {config.MIN_REVIEWS}-{config.MAX_REVIEWS} "
          f"review filter\n")

    print("3/4  Auditing websites (this is the slow part)...")
    def do_audit(p):
        return p, website_audit.audit(p.get("websiteUri", ""))

    rows = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        for i, (p, verdict) in enumerate(pool.map(do_audit, candidates), 1):
            name = p["displayName"]["text"]
            print(f"     [{i}/{len(candidates)}] {name}: {verdict['status']}")
            # Only sites with a real problem are leads.
            if verdict["status"] in ("CURRENT", "MINOR_ISSUES"):
                continue
            reviews = p.get("userRatingCount") or 0
            rows.append({
                "name": name,
                "city": city_of(p.get("formattedAddress")) or p.get("_found_near", ""),
                "status": verdict["status"].replace("_", " ").title(),
                "url": p.get("websiteUri", ""),
                "reviews": reviews,
                "rating": p.get("rating", ""),
                "phone": p.get("nationalPhoneNumber", ""),
                "address": p.get("formattedAddress", ""),
                "why": verdict["summary"],
                "maps": p.get("googleMapsUri", ""),
                "score": opportunity_score(verdict["score"], reviews),
            })

    rows.sort(key=lambda r: r["score"], reverse=True)
    rows = rows[:50]
    print(f"\n4/4  Writing {len(rows)} qualified leads to {OUT}...")
    write_workbook(rows)
    print(f"     Done -> {OUT}")

    if len(rows) < 50:
        print(f"\n     NOTE: only {len(rows)} businesses met every criterion. "
              "Widen MIN/MAX_REVIEWS or add towns in config.py for more.")


def write_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mortgage Prospects"

    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, start=1):
        ws.append([
            i, r["name"], r["city"], r["status"], r["url"], r["reviews"],
            r["rating"], r["phone"], r["address"], r["why"], r["maps"], r["score"],
        ])
        row_no = i + 1

        # Make both link columns actually clickable.
        for col, target in ((5, r["url"]), (11, r["maps"])):
            if target:
                c = ws.cell(row=row_no, column=col)
                c.hyperlink = target
                c.font = Font(color="0563C1", underline="single")

        fill = FILL_HOT if r["score"] >= 80 else FILL_WARM if r["score"] >= 55 else FILL_COOL
        ws.cell(row=row_no, column=1).fill = fill
        ws.cell(row=row_no, column=4).fill = fill
        ws.cell(row=row_no, column=10).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [8, 34, 15, 14, 42, 9, 8, 15, 46, 52, 30, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(OUT)


if __name__ == "__main__":
    try:
        main()
    except places.PlacesError as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        sys.exit(1)
